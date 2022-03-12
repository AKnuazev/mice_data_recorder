import datetime
import dearpygui.dearpygui as dpg
from math import sin, cos
import openpyxl

# Open book
mice_workbook = openpyxl.load_workbook('../data/mice_workbook.xlsx')
mice_workbook_sheetnames = mice_workbook.sheetnames


def loadLastRecord():
    with open("measurements.file") as f:
        last_line = ""
        for line in f:
            if (line != ""):
                last_line = line
        return Measurement(last_line)


class Mouse:
    def __init__(self, name):
        self.name = name
        self.start_record = Measurement()
        self.sheet = self.init_sheet()

    def init_sheet(self):
        if self.name in mice_workbook_sheetnames:
            print(f"✅ Mouse '{self.name}' found, list loaded")
            return mice_workbook[self.name]
        else:
            print(f"⚠️ Mouse '{self.name}' NOT found, new list created")
            new_sheet = mice_workbook.create_sheet(self.name)

            # Set labels
            new_sheet.cell(row=1, column=1).value = "id"
            new_sheet.cell(row=1, column=2).value = "time"
            new_sheet.cell(row=1, column=3).value = "value 1"
            new_sheet.cell(row=1, column=4).value = "value 2"
            new_sheet.cell(row=1, column=5).value = "value 3"

            # new_sheet.cell(row=1, column=7).value = "(temp) start record"
            # new_sheet.cell(row=1, column=8).value = "(temp) end record"

            mice_workbook.save("../data/mice_workbook.xlsx")
            return new_sheet

    def start_recording(self):
        self.start_record = loadLastRecord()
        print("start record:", self.start_record)
        self.sheet.cell(row=1, column=7).value = str(self.start_record)
        mice_workbook.save("../data/mice_workbook.xlsx")

    def stop_recording(self):
        self.end_record = loadLastRecord()
        print("end record:", self.end_record)
        self.sheet.cell(row=1, column=8).value = str(self.end_record)
        mice_workbook.save("../data/mice_workbook.xlsx")

        # load start record from sheet
        self.start_record = Measurement(self.sheet.cell(row=1, column=7).value)

        # Now we need do get our interval
        with open("measurements.file") as f:
            got_in_interval = False
            for line in f:
                measurement = Measurement(line)
                if measurement == self.start_record:
                    got_in_interval = True
                if got_in_interval:
                    row = str(measurement).split()
                    self.sheet.append(row)
            mice_workbook.save("../data/mice_workbook.xlsx")

    def get_measurements(self):
        measurements = []
        for row in self.sheet.iter_rows(min_row=2, min_col=1, max_col=5):
            row_str = ""
            for cell in row:
                row_str += str(cell.value) + " "
            measurements.append(Measurement(row_str))
        return measurements

    def __str__(self):
        return self.name


class Measurement:
    def __init__(self, line=None):
        self.lined = ""
        if line is None:
            self.id = None
            self.time = None
            self.value1 = None
            self.value2 = None
            self.value3 = None
        else:
            self.parse_line(line)

    def parse_line(self, line):
        self.lined = line

        line_data = line.split()
        self.id = line_data[0]
        self.time = datetime.datetime.strptime(line_data[1], '%H:%M:%S')
        self.value1 = line_data[2]
        self.value2 = line_data[3]
        self.value3 = line_data[4]

    def __eq__(self, other):
        return self.time == other.time and self.id == other.id

    def __str__(self):
        return self.lined


# Load saved mice data
mice_list = []
with open("../mice.txt") as f:
    for line in f:
        mouse = Mouse(line.split()[0])
        mice_list.append(mouse)


# Modal mouse window
# with dpg.window(label="Add mouse", modal=True, show=False, id="modal_id", no_title_bar=True):
#     dpg.add_text("All those beautiful files will be deleted.\nThis operation cannot be undone!")
#     dpg.add_separator()
#     dpg.add_checkbox(label="Don't ask me next time")
#     with dpg.group(horizontal=True):
#         dpg.add_button(label="OK", width=75, callback=lambda: dpg.configure_item("modal_id", show=False))
#         dpg.add_button(label="Cancel", width=75, callback=lambda: dpg.configure_item("modal_id", show=False))

def add_and_load_image(image_path, parent=None):
    width, height, channels, data = dpg.load_image(image_path)

    with dpg.texture_registry() as reg_id:
        texture_id = dpg.add_static_texture(width, height, data, parent=reg_id, tag="img333")
    if parent is None:
        return dpg.add_image(texture_id, width=170, height=120)
    else:
        return dpg.add_image(texture_id, parent=parent, width=100, height=70)


def getCurrMouse():
    mouse_name = dpg.get_value('mice_combo_label')
    for mouse in mice_list:
        if mouse.name == mouse_name:
            print("found mouse")
            return mouse

    print(f"❌ Mouse '{mouse_name}' NOT found")
    return None


def start_callback():
    mouse = getCurrMouse()
    if mouse:
        mouse.start_recording()
        print("started recording")


def stop_callback():
    mouse = getCurrMouse()
    if mouse:
        mouse.stop_recording()
        print("stoped recording")
    update_chart()


dpg.create_context()

time_datax = []
particles_datay = []


def update_chart():
    mouse = getCurrMouse()
    measurements = mouse.get_measurements()
    counter = 0
    for measurement in measurements:
        time_datax.append(counter)
        particles_datay.append(float(measurement.value1))
        counter += 1

    dpg.set_value('series_tag', [time_datax, particles_datay])
    dpg.set_item_label('series_tag', mouse.name)


dpg.create_context()
# with dpg.font_registry():
#     default_font = dpg.add_font("assets/Coming Sans Free Trial.otf", 20)


with dpg.window(label="Mice Data Recorder", tag="Primary Window"):
    # dpg.add_image(texture_id)
    # dpg.bind_font(default_font)

    add_and_load_image("../assets/mouse.JPG")
    dpg.add_text("Make steps:")
    dpg.add_text("")
    # dpg.show_font_manager()

    with dpg.group():
        dpg.add_text("1. Save file in old program")
        dpg.add_checkbox(label="i saved", tag="is_old_saved")
        dpg.add_text("")

    with dpg.group():
        dpg.add_text("2. Choose current mouse (you can edit mice list in mice.txt file)")
        dpg.add_combo(items=mice_list, tag="mice_combo_label", callback=update_chart)
        dpg.add_text("")

    with dpg.group():
        dpg.add_text("3. Choose action")
        dpg.add_button(label="Start", callback=start_callback)
        dpg.add_button(label="Stop", callback=stop_callback)
        dpg.add_text("")

    # create plot
    with dpg.group():
        with dpg.plot(label=dpg.get_value("Particle Number Dynamics"), height=400, width=800):
            # optionally create legend
            dpg.add_plot_legend()

            # REQUIRED: create x and y axes
            dpg.add_plot_axis(dpg.mvXAxis, label="time")
            dpg.add_plot_axis(dpg.mvYAxis, label="particles", tag="y_axis")

            # series belong to a y axis
            dpg.add_line_series(time_datax, particles_datay, label="current mouse", parent="y_axis", tag="series_tag")

dpg.create_viewport(title='Mice Data Recorder')  # , width=600, height=200
dpg.setup_dearpygui()
dpg.show_viewport()
dpg.set_primary_window("Primary Window", True)

dpg.start_dearpygui()
dpg.destroy_context()
